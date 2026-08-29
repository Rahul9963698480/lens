"""Convert an uploaded XLSX workbook into a local DuckDB database."""

from __future__ import annotations

import math
import re
from datetime import date, datetime
from numbers import Integral, Real
from pathlib import Path
from typing import Any

import duckdb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings

_NON_ALNUM = re.compile(r"[^a-z0-9_]+")
_MULTI_UNDERSCORE = re.compile(r"_+")


class XlsxIngestError(ValueError):
    """Workbook could not be parsed into DuckDB tables."""


def sanitize_identifier(name: str, *, fallback: str = "col") -> str:
    text = str(name or "").strip().lower()
    text = _NON_ALNUM.sub("_", text)
    text = _MULTI_UNDERSCORE.sub("_", text).strip("_")
    if not text:
        text = fallback
    if text[0].isdigit():
        text = f"{fallback}_{text}"
    return text


def unique_identifier(base: str, used: set[str]) -> str:
    if base not in used:
        used.add(base)
        return base
    index = 2
    while f"{base}_{index}" in used:
        index += 1
    name = f"{base}_{index}"
    used.add(name)
    return name


def _quote_ident(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def _is_null(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _is_integer_like(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, Integral):
        return True
    if isinstance(value, Real):
        number = float(value)
        return math.isfinite(number) and number.is_integer()
    return False


def _is_numeric(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, Integral):
        return True
    if isinstance(value, Real):
        return math.isfinite(float(value))
    return False


def _is_date_like(value: Any) -> bool:
    return isinstance(value, (datetime, date, pd.Timestamp))


def infer_column_type(series: pd.Series) -> str:
    values = [v for v in series.tolist() if not _is_null(v)]
    if not values:
        return "VARCHAR"
    if all(_is_integer_like(v) for v in values):
        return "BIGINT"
    if all(_is_numeric(v) for v in values):
        return "DOUBLE"
    if all(_is_date_like(v) for v in values):
        return "VARCHAR"
    return "VARCHAR"


def _coerce_frame_value(value: Any, column_type: str) -> Any:
    if _is_null(value):
        return None
    if column_type == "BIGINT":
        return int(value)
    if column_type == "DOUBLE":
        return float(value)
    if _is_date_like(value):
        ts = pd.Timestamp(value)
        if ts.hour == 0 and ts.minute == 0 and ts.second == 0 and ts.microsecond == 0:
            return ts.date().isoformat()
        return ts.isoformat(timespec="seconds")
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _sanitize_headers(raw_headers: tuple[Any, ...]) -> list[str]:
    used_cols: set[str] = set()
    columns: list[str] = []
    for raw in raw_headers:
        base = sanitize_identifier(str(raw), fallback="col")
        columns.append(unique_identifier(base, used_cols))
    return columns


def _normalize_row(row: tuple[Any, ...], width: int) -> list[Any]:
    values = list(row[:width])
    if len(values) < width:
        values.extend([None] * (width - len(values)))
    return values


def _infer_types(columns: list[str], rows: list[list[Any]]) -> list[str]:
    df = pd.DataFrame(rows, columns=columns)
    df = df.dropna(axis=0, how="all")
    if df.empty:
        raise XlsxIngestError("Sheet has no data rows.")
    return [infer_column_type(df[col]) for col in columns]


def _rows_to_frame(
    rows: list[list[Any]], columns: list[str], types: list[str]
) -> pd.DataFrame:
    coerced = [
        [_coerce_frame_value(value, col_type) for value, col_type in zip(row, types)]
        for row in rows
    ]
    return pd.DataFrame(coerced, columns=columns)


def _configure_connection(conn: duckdb.DuckDBPyConnection) -> None:
    if settings.DUCKDB_MEMORY_LIMIT:
        conn.execute(f"SET memory_limit = '{settings.DUCKDB_MEMORY_LIMIT}'")
    if settings.DUCKDB_THREADS > 0:
        conn.execute(f"SET threads = {int(settings.DUCKDB_THREADS)}")


def _insert_frame_chunked(
    conn: duckdb.DuckDBPyConnection,
    table: str,
    df: pd.DataFrame,
    *,
    chunk_size: int,
) -> None:
    if df.empty:
        return
    quoted_cols = ", ".join(_quote_ident(c) for c in df.columns)
    for start in range(0, len(df), chunk_size):
        chunk = df.iloc[start : start + chunk_size]
        conn.register("_ingest_chunk", chunk)
        try:
            conn.execute(
                f"INSERT INTO {_quote_ident(table)} ({quoted_cols}) "
                f"SELECT {quoted_cols} FROM _ingest_chunk"
            )
        finally:
            conn.unregister("_ingest_chunk")


def _ingest_sheet_streaming(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    *,
    sheet_name: str,
    used_tables: set[str],
    chunk_size: int,
) -> str | None:
    row_iter = ws.iter_rows(values_only=True)
    header = next(row_iter, None)
    if header is None or all(_is_null(value) for value in header):
        return None

    columns = _sanitize_headers(header)
    if not columns:
        return None

    width = len(columns)
    table = unique_identifier(
        sanitize_identifier(str(sheet_name), fallback="sheet"),
        used_tables,
    )
    types: list[str] | None = None
    pending: list[list[Any]] = []

    def _create_table(sample_rows: list[list[Any]]) -> None:
        nonlocal types
        types = _infer_types(columns, sample_rows)
        col_defs = ", ".join(
            f"{_quote_ident(col)} {col_type}" for col, col_type in zip(columns, types)
        )
        conn.execute(f"CREATE TABLE {_quote_ident(table)} ({col_defs})")

    def _flush(rows: list[list[Any]]) -> None:
        if not rows:
            return
        trimmed = [row[: len(columns)] for row in rows]
        frame = _rows_to_frame(trimmed, columns, types or [])
        _insert_frame_chunked(conn, table, frame, chunk_size=chunk_size)

    for row in row_iter:
        values = _normalize_row(row, width)
        if all(_is_null(value) for value in values):
            continue

        if types is None:
            pending.append(values)
            if len(pending) >= chunk_size:
                _create_table(pending)
                _flush(pending)
                pending = []
            continue

        pending.append(values)
        if len(pending) >= chunk_size:
            _flush(pending)
            pending = []

    if types is None:
        if not pending:
            return None
        _create_table(pending)
        _flush(pending)
        return table

    _flush(pending)
    return table


def workbook_to_duckdb(xlsx_path: str | Path, duckdb_path: str | Path) -> list[str]:
    """Parse an .xlsx file into DuckDB tables. Returns created table names."""
    path = Path(xlsx_path)
    if path.suffix.lower() != ".xlsx":
        raise XlsxIngestError("Only .xlsx files are supported.")

    duckdb_path = Path(duckdb_path)
    duckdb_path.parent.mkdir(parents=True, exist_ok=True)
    if duckdb_path.exists():
        duckdb_path.unlink()

    used_tables: set[str] = set()
    created: list[str] = []
    chunk_size = settings.XLSX_INGEST_CHUNK_SIZE

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise XlsxIngestError(
            f"Could not parse the spreadsheet (corrupt or unsupported format): {exc}"
        ) from exc

    conn = duckdb.connect(str(duckdb_path))
    _configure_connection(conn)
    try:
        if not workbook.sheetnames:
            raise XlsxIngestError("Workbook is empty — no sheets found.")

        for sheet_name in workbook.sheetnames:
            try:
                ws = workbook[sheet_name]
                table = _ingest_sheet_streaming(
                    ws,
                    conn,
                    sheet_name=str(sheet_name),
                    used_tables=used_tables,
                    chunk_size=chunk_size,
                )
            except Exception:
                continue
            if table:
                created.append(table)
    except Exception:
        if duckdb_path.exists():
            duckdb_path.unlink()
        raise
    finally:
        conn.close()
        workbook.close()

    if not created:
        if duckdb_path.exists():
            duckdb_path.unlink()
        raise XlsxIngestError("Workbook is empty — no sheets with data.")

    return created
